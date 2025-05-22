from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
import sqlite3
from datetime import datetime
from io import BytesIO
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Для управления сессиями
DATABASE = 'database.db'

# Подключение к базе данных
def get_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    return db

# Создание таблиц и начальных данных
def create_tables():
    db = get_db()
    db.execute('''CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    role TEXT NOT NULL CHECK(role IN ('operator', 'admin'))
                  )''')
    db.execute('''CREATE TABLE IF NOT EXISTS categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL
                  )''')
    db.execute('''CREATE TABLE IF NOT EXISTS stalls (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category_id INTEGER NOT NULL,
                    stall_number TEXT NOT NULL,
                    FOREIGN KEY (category_id) REFERENCES categories (id)
                  )''')
    db.execute('''CREATE TABLE IF NOT EXISTS assignments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    stall_id INTEGER NOT NULL,
                    operator_id INTEGER NOT NULL,
                    assignee_name TEXT NOT NULL,
                    assignee_surname TEXT NOT NULL,
                    amount REAL NOT NULL,
                    assigned_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (stall_id) REFERENCES stalls (id),
                    FOREIGN KEY (operator_id) REFERENCES users (id)
                  )''')
    db.commit()

    # Инициализация начальных данных
    if db.execute('SELECT COUNT(*) FROM categories').fetchone()[0] == 0:
        db.execute('INSERT INTO categories (name) VALUES (?)', ('Продукты',))
        db.execute('INSERT INTO categories (name) VALUES (?)', ('Скот',))
        db.execute('INSERT INTO categories (name) VALUES (?)', ('Одежда',))
    if db.execute('SELECT COUNT(*) FROM stalls').fetchone()[0] == 0:
        categories = db.execute('SELECT id FROM categories').fetchall()
        for cat in categories:
            for i in range(1, 4):
                db.execute('INSERT INTO stalls (category_id, stall_number) VALUES (?, ?)', (cat['id'], f'A{i}'))
    if db.execute('SELECT COUNT(*) FROM users').fetchone()[0] == 0:
        for i in range(1, 6):
            db.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)', (f'admin{i}', 'adminpass', 'admin'))
        db.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)', ('operator1', 'oppass1', 'operator'))
        db.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)', ('operator2', 'oppass2', 'operator'))
    db.commit()

# Маршрут для входа
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password)).fetchone()
        if user:
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['username'] = user['username']
            if user['role'] == 'operator':
                return redirect(url_for('operator_dashboard'))
            else:
                return redirect(url_for('admin_reports'))
        else:
            return 'Неверные учетные данные'
    return render_template('login.html')

# Маршрут для выхода
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('role', None)
    session.pop('username', None)
    return redirect(url_for('login'))

# Панель оператора
@app.route('/operator_dashboard')
def operator_dashboard():
    if 'user_id' not in session or session['role'] != 'operator':
        return redirect(url_for('login'))
    db = get_db()
    today = datetime.now().date().isoformat()
    assignments = db.execute('''
        SELECT a.*, c.name as category_name, s.stall_number
        FROM assignments a
        JOIN stalls s ON a.stall_id = s.id
        JOIN categories c ON s.category_id = c.id
        WHERE DATE(a.assigned_at) = ?
    ''', (today,)).fetchall()
    return render_template('operator_dashboard.html', assignments=assignments)

# Форма для назначения места
@app.route('/assign_place', methods=['GET'])
def assign_place():
    if 'user_id' not in session or session['role'] != 'operator':
        return redirect(url_for('login'))
    db = get_db()
    categories = db.execute('SELECT * FROM categories').fetchall()
    return render_template('assign_place.html', categories=categories)

# API для получения доступных стендов
@app.route('/get_available_stalls/<int:category_id>')
def get_available_stalls(category_id):
    db = get_db()
    today = datetime.now().date().isoformat()
    all_stalls = db.execute('SELECT id, stall_number FROM stalls WHERE category_id = ?', (category_id,)).fetchall()
    assigned_stalls = db.execute('''
        SELECT s.id
        FROM stalls s
        JOIN assignments a ON s.id = a.stall_id
        WHERE s.category_id = ? AND DATE(a.assigned_at) = ?
    ''', (category_id, today)).fetchall()
    assigned_ids = [stall['id'] for stall in assigned_stalls]
    available_stalls = [stall for stall in all_stalls if stall['id'] not in assigned_ids]
    return jsonify([{'id': stall['id'], 'number': stall['stall_number']} for stall in available_stalls])

# Обработка назначения места
@app.route('/assign', methods=['POST'])
def assign():
    if 'user_id' not in session or session['role'] != 'operator':
        return redirect(url_for('login'))
    category_id = request.form['category_id']
    assignee_name = request.form['assignee_name']
    assignee_surname = request.form['assignee_surname']
    amount = request.form['amount']
    operator_id = session['user_id']
    db = get_db()
    db.execute('INSERT INTO assignments (category_id, operator_id, assignee_name, assignee_surname, amount) VALUES (?, ?, ?, ?, ?)',
               (category_id, operator_id, assignee_name, assignee_surname, amount))
    db.commit()
    assignment_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
    return redirect(url_for('receipt', assignment_id=assignment_id))

# Печать чека
@app.route('/receipt/<int:assignment_id>')
def receipt(assignment_id):
    db = get_db()
    assignment = db.execute('''
        SELECT a.*, c.name as category_name, s.stall_number, u.username as operator_name
        FROM assignments a
        JOIN stalls s ON a.stall_id = s.id
        JOIN categories c ON s.category_id = c.id
        JOIN users u ON a.operator_id = u.id
        WHERE a.id = ?
    ''', (assignment_id,)).fetchone()
    if not assignment:
        return 'Назначение не найдено', 404
    return render_template('receipt.html', assignment=assignment)

# Отчеты администратора
@app.route('/admin_reports', methods=['GET', 'POST'])
def admin_reports():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        db = get_db()
        assignments = db.execute('''
            SELECT a.*, c.name as category_name, s.stall_number, u.username as operator_name
            FROM assignments a
            JOIN stalls s ON a.stall_id = s.id
            JOIN categories c ON s.category_id = c.id
            JOIN users u ON a.operator_id = u.id
            WHERE a.assigned_at BETWEEN ? AND ?
        ''', (start_date, end_date)).fetchall()
        total_assignments = len(assignments)
        total_amount = sum(assignment['amount'] for assignment in assignments)
        return render_template('admin_reports.html', assignments=assignments, start_date=start_date, end_date=end_date, total_assignments=total_assignments, total_amount=total_amount)
    return render_template('admin_reports.html')

# Excel-отчет для администраторов
@app.route('/admin_reports_excel')
def admin_reports_excel():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    db = get_db()
    assignments = db.execute('''
        SELECT a.*, c.name as category_name, s.stall_number, u.username as operator_name
        FROM assignments a
        JOIN stalls s ON a.stall_id = s.id
        JOIN categories c ON s.category_id = c.id
        JOIN users u ON a.operator_id = u.id
        WHERE a.assigned_at BETWEEN ? AND ?
    ''', (start_date, end_date)).fetchall()

    # Создание Excel-файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по назначениям"

    # Добавление заголовков
    headers = ['Дата', 'Категория', 'Стенд', 'Арендатор (имя)', 'Арендатор (фамилия)', 'Сумма', 'Оператор']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Добавление данных
    for row_num, assignment in enumerate(assignments, 2):
        ws.cell(row=row_num, column=1, value=assignment['assigned_at'])
        ws.cell(row=row_num, column=2, value=assignment['category_name'])
        ws.cell(row=row_num, column=3, value=assignment['stall_number'])
        ws.cell(row=row_num, column=4, value=assignment['assignee_name'])
        ws.cell(row=row_num, column=5, value=assignment['assignee_surname'])
        ws.cell(row=row_num, column=6, value=assignment['amount'])
        ws.cell(row=row_num, column=7, value=assignment['operator_name'])

    # Сохранение в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Отправка файла
    return send_file(output, download_name='assignments_report.xlsx', as_attachment=True)

# Excel-отчет для операторов
@app.route('/operator_daily_report')
def operator_daily_report():
    if 'user_id' not in session or session['role'] != 'operator':
        return redirect(url_for('login'))
    today = datetime.now().date().isoformat()
    db = get_db()
    assignments = db.execute('''
        SELECT a.*, c.name as category_name, s.stall_number
        FROM assignments a
        JOIN stalls s ON a.stall_id = s.id
        JOIN categories c ON s.category_id = c.id
        WHERE a.operator_id = ? AND DATE(a.assigned_at) = ?
    ''', (session['user_id'], today)).fetchall()

    # Создание Excel-файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сегодняшние назначения"

    # Добавление заголовков
    headers = ['Дата', 'Категория', 'Стенд', 'Арендатор (имя)', 'Арендатор (фамилия)', 'Сумма']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Добавление данных
    for row_num, assignment in enumerate(assignments, 2):
        ws.cell(row=row_num, column=1, value=assignment['assigned_at'])
        ws.cell(row=row_num, column=2, value=assignment['category_name'])
        ws.cell(row=row_num, column=3, value=assignment['stall_number'])
        ws.cell(row=row_num, column=4, value=assignment['assignee_name'])
        ws.cell(row=row_num, column=5, value=assignment['assignee_surname'])
        ws.cell(row=row_num, column=6, value=assignment['amount'])

    # Сохранение в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Отправка файла
    return send_file(output, attachment_filename='todays_assignments.xlsx', as_attachment=True)

@app.route('/')
def index():
    return redirect(url_for('login'))  # Redirect to login page, for example


if __name__ == '__main__':
    create_tables()
    app.run(debug=True)